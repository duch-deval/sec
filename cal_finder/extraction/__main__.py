#!/usr/bin/env python3
import csv
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from .extract_fields import run_pipeline

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def parse_date_range(arg: str):
    start_s, end_s = arg.split("..")
    start = datetime.strptime(start_s.strip(), "%Y-%m-%d")
    end = datetime.strptime(end_s.strip(), "%Y-%m-%d")
    dates = []
    cur = start
    while cur <= end:
        dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    label = f"{start_s.strip()}..{end_s.strip()}"
    return dates, label


def merge_csvs_to_xlsx(date_csvs, output_name: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for date_str, csv_path in date_csvs:
        ws = wb.create_sheet(title=date_str)
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

    if not wb.sheetnames:
        return None

    wb.save(output_name)
    return output_name


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python -m extraction <YYYY-MM-DD>                    # single date")
        print("  python -m extraction <YYYY-MM-DD>..<YYYY-MM-DD>      # date range -> merged xlsx")
        print("  python -m extraction <file.xlsx|file.csv>             # sample file mode")
        sys.exit(1)

    arg = sys.argv[1]
    mapping = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if arg.endswith((".xlsx", ".csv")):
        input_path = Path(arg)
        if not input_path.exists():
            sys.exit(f"File not found: {input_path}")

        output_dir = Path(input_path.stem)
        output_dir.mkdir(parents=True, exist_ok=True)

        from .output_writer import run as download_samples
        download_samples(input_path, output_dir)

        success = run_pipeline(output_dir, mapping, verbose=True)
        sys.exit(0 if success else 1)

    elif ".." in arg:
        dates, range_label = parse_date_range(arg)
        logger.info("Date range: %s (%d calendar days)", range_label, len(dates))

        date_csvs = []
        all_success = True

        for date in dates:
            root = Path(date)
            if not root.exists():
                continue
            logger.info("Processing %s", date)
            ok = run_pipeline(root, mapping, verbose=True)
            if not ok:
                all_success = False

            csv_path = root / f"{date}.csv"
            if csv_path.exists() and csv_path.stat().st_size > 0:
                date_csvs.append((date, csv_path))

        if date_csvs:
            out = merge_csvs_to_xlsx(date_csvs, f"{range_label}.xlsx")
            if out:
                logger.info("Merged %d dates -> %s", len(date_csvs), out)
            else:
                logger.warning("No data to merge")
        else:
            logger.warning("No extraction CSVs produced for range %s", range_label)

        sys.exit(0 if all_success else 1)

    else:
        root = Path(arg)
        success = run_pipeline(root, mapping, verbose=True)
        sys.exit(0 if success else 1)