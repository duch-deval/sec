import os
from openpyxl import load_workbook

def main():
    candidates = [
        os.path.join("cal_finder", "baseline", "bin", "baseline.xlsx"),
        os.path.join("baseline", "bin", "baseline.xlsx"),
        os.path.join("bin", "baseline.xlsx")
    ]

    path = None
    for p in candidates:
        if os.path.exists(p):
            if os.path.isfile(p):
                path = p
                break
            else:
                print(f"Path exists but is not a file: {p}")

    if path is None:
        print("baseline.xlsx not found in baseline\\bin or bin")
        return

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=4).value
        if v is None:
            continue
        s = str(v).strip()
        if not s or s == "File Link":
            continue

        if s.startswith("https://www.sec.gov/ix?doc=/Archives/"):
            s = s.replace("https://www.sec.gov/ix?doc=/Archives", "https://www.sec.gov/Archives", 1)

        if s.startswith("https://www.sec.gov/Archives/"):
            print(s)

if __name__ == "__main__":
    main()
